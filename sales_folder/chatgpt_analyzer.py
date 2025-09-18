import os
import pandas as pd
from openai import OpenAI
from dotenv import load_dotenv
from openpyxl import load_workbook
from typing import Optional, Dict, Any, List

# Добавляем путь к родительской директории
import sys
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from utils import load_excel_with_cache, find_column, clean_numeric_data

# Загружаем переменные окружения
load_dotenv()

class ChatGPTAnalyzer:
    def __init__(self):
        """Инициализация ChatGPT AI анализатора"""
        self.api_key = os.getenv("OPENAI_API_KEY")
        self.client = None
        
        if self.api_key:
            # Настройка OpenAI только если API ключ доступен
            self.client = OpenAI(api_key=self.api_key)
    
    def analyze_sales_data(self, excel_file_path: str, focus: Optional[str] = None) -> str:
        """
        Анализирует данные продаж из Excel файла с помощью ChatGPT AI
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            focus (str, optional): Фокус анализа (например, "top_sales")
            
        Returns:
            str: Анализ данных в текстовом формате
        """
        if not self.client:
            return "❌ ChatGPT AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            # Читаем данные из Excel с использованием кэша
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            salary_df = load_excel_with_cache(excel_file_path, 'зарплата')
            
            # Подготавливаем данные для анализа
            analysis_data = self._prepare_data_for_analysis(sales_df, salary_df)
            
            # Создаем промпт для ChatGPT
            prompt = self._create_analysis_prompt(analysis_data, focus)
            
            # Получаем анализ от ChatGPT
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Ты эксперт по анализу данных продаж и финансов. Отвечай на русском языке, давай конкретные рекомендации и инсайты."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.7
            )
            
            return response.choices[0].message.content
            
        except Exception as e:
            return f"Ошибка при анализе данных: {str(e)}"
    
    def _prepare_data_for_analysis(self, sales_df: pd.DataFrame, salary_df: pd.DataFrame) -> Dict[str, Any]:
        """Подготавливает данные для анализа"""
        
        # Определяем колонку с ценами/суммами
        price_column = None
        possible_price_columns = ['price', 'сумма', 'стоимость', 'total', 'amount']
        
        for col in possible_price_columns:
            if col in sales_df.columns:
                price_column = col
                break
        
        # Определяем колонку с основным менеджером
        manager_column = None
        possible_manager_columns = ['manager', 'менеджер']
        
        for col in possible_manager_columns:
            if col in sales_df.columns:
                manager_column = col
                break
        
        # Определяем колонку с ROP
        rop_column = None
        possible_rop_columns = ['employee ROP', 'rop', 'РОП']
        
        for col in possible_rop_columns:
            if col in sales_df.columns:
                rop_column = col
                break
        
        # Основная статистика по продажам
        sales_stats = {
            'total_sales': pd.to_numeric(sales_df[price_column], errors='coerce').sum() if price_column else 0,
            'avg_sale': pd.to_numeric(sales_df[price_column], errors='coerce').mean() if price_column else 0,
            'total_transactions': len(sales_df),
            'unique_managers': sales_df[manager_column].nunique() if manager_column else 0,
            'unique_rop': sales_df[rop_column].nunique() if rop_column else 0,
            'price_column_used': price_column,
            'manager_column_used': manager_column,
            'rop_column_used': rop_column
        }
        
        # Статистика по менеджерам
        if manager_column and price_column:
            try:
                # Убираем NaN значения для корректного группирования
                valid_data = sales_df.dropna(subset=[manager_column])
                valid_data[price_column] = pd.to_numeric(valid_data[price_column], errors='coerce')
                manager_stats = valid_data.groupby(manager_column)[price_column].agg(['sum', 'count', 'mean']).round(2)
                sales_stats['manager_performance'] = manager_stats.to_dict('index')
            except Exception as e:
                sales_stats['manager_performance'] = {}
        
        # Статистика по ROP сотрудникам
        if rop_column and price_column:
            try:
                valid_data = sales_df.dropna(subset=[rop_column])
                valid_data[price_column] = pd.to_numeric(valid_data[price_column], errors='coerce')
                rop_stats = valid_data.groupby(rop_column)[price_column].agg(['sum', 'count', 'mean']).round(2)
                sales_stats['rop_performance'] = rop_stats.to_dict('index')
            except Exception as e:
                sales_stats['rop_performance'] = {}
        
        # Статистика по зарплатам (если есть лист зарплата)
        salary_stats = {'note': 'Данные по зарплатам не анализируются для листа продаж'}
        if salary_df is not None and not salary_df.empty:
            # Определяем колонки зарплат
            salary_column = None
            bonus_columns = []
            
            for col in salary_df.columns:
                if 'оклад' in col.lower() or 'salary' in col.lower():
                    salary_column = col
                elif 'бонус' in col.lower() or 'bonus' in col.lower():
                    bonus_columns.append(col)
            
            if salary_column:
                salary_stats = {
                    'total_salary': pd.to_numeric(salary_df[salary_column], errors='coerce').sum(),
                    'avg_salary': pd.to_numeric(salary_df[salary_column], errors='coerce').mean(),
                    'total_bonuses': sum(pd.to_numeric(salary_df[col], errors='coerce').sum() for col in bonus_columns),
                    'salary_column_used': salary_column,
                    'bonus_columns_used': bonus_columns
                }
        
        return {
            'sales': sales_stats,
            'salary': salary_stats,
            'raw_sales_data': sales_df.head(10).to_dict('records'),  # Первые 10 записей
            'raw_salary_data': salary_df.head(10).to_dict('records') if salary_df is not None else []
        }
    
    def _create_analysis_prompt(self, data: Dict[str, Any], focus: Optional[str] = None) -> str:
        """Создает промпт для анализа данных"""
        # Базовый промпт
        base_prompt = f"""
        Проанализируй следующие данные о продажах и зарплатах компании:

        ОСНОВНАЯ СТАТИСТИКА ПРОДАЖ:
        - Общая сумма продаж: {data['sales']['total_sales']:,.0f} тенге
        - Средняя сумма продажи: {data['sales']['avg_sale']:,.0f} тенге
        - Общее количество транзакций: {data['sales']['total_transactions']}
        - Количество менеджеров: {data['sales']['unique_managers']}
        - Количество ROP сотрудников: {data['sales']['unique_rop']}

        СТАТИСТИКА ПО ЗАРПЛАТАМ:
        {self._format_salary_stats(data['salary'])}

        ПЕРВЫЕ 10 ЗАПИСЕЙ ПРОДАЖ:
        {data['raw_sales_data']}

        ПЕРВЫЕ 10 ЗАПИСЕЙ ЗАРПЛАТ:
        {data['raw_salary_data']}
        """
        
        # Добавляем специфичные инструкции в зависимости от фокуса
        if focus == "top_sales":
            prompt = base_prompt + """
            
            СФОКУСИРУЙСЯ НА ТОП ПРОДАЖАХ:
            - Проанализируй лучшие продажи и их характеристики
            - Выдели топ-менеджеров по продажам
            - Определи факторы успеха в продажах
            - Дай рекомендации по улучшению продаж

            Проведи КРАТКИЙ анализ и предоставь:
            1. Топ-3 менеджера по продажам
            2. Ключевые факторы успеха
            3. Главный вывод и 1-2 рекомендации по улучшению продаж

            ВАЖНО: Отвечай кратко, без лишних объяснений. Максимум 5-7 предложений.
            """
        else:
            prompt = base_prompt + """
            
            Проведи КРАТКИЙ анализ и предоставь:
            1. Ключевые показатели (2-3 главные цифры)
            2. Лучшие и худшие результаты
            3. Главный вывод и 1-2 рекомендации

            ВАЖНО: Отвечай кратко, без лишних объяснений. Максимум 5-7 предложений.
            """
        
        return prompt
    
    def _format_salary_stats(self, salary_data: Dict[str, Any]) -> str:
        """Форматирует статистику по зарплатам"""
        if isinstance(salary_data, dict) and 'note' in salary_data:
            return salary_data['note']
        
        if isinstance(salary_data, dict):
            result = ""
            if 'total_salary' in salary_data:
                result += f"- Общая сумма окладов: {salary_data['total_salary']:,.0f} тенге\n"
            if 'avg_salary' in salary_data:
                result += f"- Средний оклад: {salary_data['avg_salary']:,.0f} тенге\n"
            if 'total_bonuses' in salary_data:
                result += f"- Общая сумма бонусов: {salary_data['total_bonuses']:,.0f} тенге\n"
            
            if not result:
                return "- Данные по зарплатам недоступны для анализа"
            
            return result.strip()
        
        return "- Данные по зарплатам недоступны"
    
    def get_quick_insights(self, excel_file_path: str) -> str:
        """
        Получает быстрые инсайты по данным
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            
        Returns:
            str: Краткие инсайты
        """
        try:
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            
            # Определяем колонку с ценами/суммами
            price_column = None
            possible_price_columns = ['price', 'сумма', 'стоимость', 'total', 'amount']
            
            for col in possible_price_columns:
                if col in sales_df.columns:
                    price_column = col
                    break
            
            if price_column is None:
                return f"❌ Не найдена колонка с ценами. Доступные колонки: {list(sales_df.columns)}"
            
            # Быстрые расчеты
            # Убираем NaN значения и преобразуем в числовой формат
            price_data = pd.to_numeric(sales_df[price_column], errors='coerce').dropna()
            
            total_sales = price_data.sum()
            avg_sale = price_data.mean()
            max_sale = price_data.max()
            min_sale = price_data.min()
            
            # Дополнительная информация по товарам
            product_info = ""
            if 'name_boiler' in sales_df.columns:
                product_counts = sales_df['name_boiler'].value_counts().head(3)
                product_info = f"\n\n🏆 ТОП-3 ТОВАРА:\n"
                for i, (product, count) in enumerate(product_counts.items(), 1):
                    product_info += f"• {i}. {product}: {count} шт.\n"
            
            # Информация по менеджерам
            manager_info = ""
            manager_columns = [col for col in sales_df.columns if 'manager' in col.lower()]
            if manager_columns:
                # Берем основной столбец менеджера
                main_manager_col = 'manager' if 'manager' in sales_df.columns else manager_columns[0]
                manager_counts = sales_df[main_manager_col].value_counts().head(3)
                manager_info = f"\n\n👥 ТОП-3 МЕНЕДЖЕРА:\n"
                for i, (manager, count) in enumerate(manager_counts.items(), 1):
                    if pd.notna(manager) and str(manager) != 'nan':
                        manager_info += f"• {i}. {manager}: {count} заказов\n"
            
            insights = f"""
📊 БЫСТРЫЕ ИНСАЙТЫ:
• Общая сумма продаж: {total_sales:,.0f} тенге
• Средняя сумма продажи: {avg_sale:,.0f} тенге
• Максимальная продажа: {max_sale:,.0f} тенге
• Минимальная продажа: {min_sale:,.0f} тенге
• Количество транзакций: {len(sales_df)}
• Колонка с ценами: {price_column}{product_info}{manager_info}
            """
            
            return insights.strip()
            
        except Exception as e:
            return f"Ошибка при получении инсайтов: {str(e)}"
    
    def generate_period_report(self, excel_file_path: str, period_request: str) -> str:
        """
        Генерирует отчет за определенный период
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            period_request (str): Запрос на отчет (например: "отчет за эту неделю")
            
        Returns:
            str: Отчет за период
        """
        if not self.client:
            return "❌ ChatGPT AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            # Читаем данные из Excel
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            
            # Создаем специальный промпт для отчетов
            prompt = f"""
            Пользователь запросил: "{period_request}"
            
            Проанализируй данные о продажах и создай подробный отчет.
            
            ДАННЫЕ О ПРОДАЖАХ:
            Колонки в данных: {list(sales_df.columns)}
            Количество записей: {len(sales_df)}
            
            ПЕРВЫЕ 10 ЗАПИСЕЙ:
            {sales_df.head(10).to_dict('records')}
            
            ПОСЛЕДНИЕ 5 ЗАПИСЕЙ:
            {sales_df.tail(5).to_dict('records')}
            
            СТАТИСТИКА:
            - Общее количество заказов: {len(sales_df)}
            """
            
            # Добавляем статистику по ценам, если есть
            price_column = find_column(sales_df, 'price')
            if price_column:
                price_data = clean_numeric_data(sales_df, price_column)
                prompt += f"""
            - Общая сумма продаж: {price_data.sum():,.0f} тенге
            - Средняя сумма заказа: {price_data.mean():,.0f} тенге
            - Максимальная продажа: {price_data.max():,.0f} тенге
            - Минимальная продажа: {price_data.min():,.0f} тенге
                """
            
            # Добавляем информацию по менеджерам
            manager_column = find_column(sales_df, 'manager')
            if manager_column:
                manager_stats = sales_df[manager_column].value_counts().head(5)
                prompt += f"""
            
            ТОП-5 МЕНЕДЖЕРОВ ПО КОЛИЧЕСТВУ ЗАКАЗОВ:
            {manager_stats.to_dict()}
                """
            
            # Добавляем информацию по товарам
            product_column = find_column(sales_df, 'product')
            if product_column:
                product_stats = sales_df[product_column].value_counts().head(5)
                prompt += f"""
            
            ТОП-5 ТОВАРОВ ПО КОЛИЧЕСТВУ ЗАКАЗОВ:
            {product_stats.to_dict()}
                """
            
            prompt += f"""
            
            ЗАДАЧА:
            Создай КРАТКИЙ отчет:
            1. Основные цифры (продажи, заказы, средний чек)
            2. ТОП-3 менеджера и товара
            3. Главный вывод
            
            ВАЖНО: 
            - Максимум 6-8 предложений
            - Конкретные цифры, минимум текста
            - Если нет данных за период - скажи сразу
            """
            
            # Получаем анализ от ChatGPT
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Ты эксперт по анализу данных продаж и финансов. Отвечай на русском языке, давай конкретные рекомендации и инсайты."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1500,
                temperature=0.7
            )
            
            return response.choices[0].message.content
            
        except Exception as e:
            return f"Ошибка при генерации отчета: {str(e)}"
    
    def universal_query(self, excel_file_path: str, user_query: str) -> str:
        """
        Универсальный обработчик запросов - может ответить на любой вопрос по данным
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            user_query (str): Любой запрос пользователя
            
        Returns:
            str: Ответ на запрос
        """
        # Проверяем, является ли запрос командой редактирования
        if self._is_edit_command(user_query):
            return self.edit_excel_data(excel_file_path, user_query)
        
        # Если это обычный запрос - используем стандартный анализ
        if not self.client:
            return "❌ ChatGPT AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            # Читаем данные из Excel
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            
            # Создаем универсальный промпт
            prompt = f"""
            Ты - умный аналитик данных. Пользователь задал вопрос о данных продаж.
            
            ЗАПРОС ПОЛЬЗОВАТЕЛЯ: "{user_query}"
            
            СПЕЦИАЛЬНЫЕ ПРАВИЛА:
            - Если спрашивают про "средний чек" - используй колонку 'price' как основную сумму заказа
            - Если спрашивают про "прибыль" - используй колонку 'purchase' 
            - Если нужны даты, а колонки 'date' нет - скажи сразу "Нет данных по датам"
            
            ДАННЫЕ О ПРОДАЖАХ:
            Структура данных (колонки): {list(sales_df.columns)}
            Общее количество записей: {len(sales_df)}
            
            ОБРАЗЕЦ ДАННЫХ (первые 5 записей):
            {sales_df.head(5).to_dict('records')}
            
            ПОСЛЕДНИЕ 3 ЗАПИСИ:
            {sales_df.tail(3).to_dict('records')}
            
            СТАТИСТИЧЕСКАЯ ИНФОРМАЦИЯ:
            """
            
            # Добавляем базовую статистику для всех числовых колонок
            numeric_columns = sales_df.select_dtypes(include=['number']).columns
            for col in numeric_columns:
                if not sales_df[col].isna().all():
                    prompt += f"""
            {col}:
              - Минимум: {sales_df[col].min()}
              - Максимум: {sales_df[col].max()}
              - Среднее: {sales_df[col].mean():.2f}
              - Сумма: {sales_df[col].sum()}
                    """
            
            # Добавляем информацию по категориальным колонкам
            categorical_columns = sales_df.select_dtypes(include=['object', 'string']).columns
            for col in categorical_columns[:5]:  # Ограничиваем до 5 колонок
                if not sales_df[col].isna().all():
                    unique_count = sales_df[col].nunique()
                    if unique_count <= 20:  # Показываем только если уникальных значений не много
                        value_counts = sales_df[col].value_counts().head(5)
                        prompt += f"""
            {col} (уникальных значений: {unique_count}):
              Топ-5 значений: {value_counts.to_dict()}
                        """
                    else:
                        prompt += f"""
            {col}: {unique_count} уникальных значений
                        """
            
            prompt += f"""
            
            ТВОЯ ЗАДАЧА:
            1. Проанализируй запрос пользователя и пойми, что именно он хочет узнать
            2. Используй предоставленные данные для ответа
            3. Если нужно сделать вычисления - сделай их
            4. Если нужно найти конкретные записи - найди их
            5. Если нужно сравнить или проанализировать тренды - проанализируй
            6. Если данных недостаточно для полного ответа - скажи об этом честно
            
            ТИПЫ ЗАПРОСОВ, КОТОРЫЕ ТЫ МОЖЕШЬ ОБРАБОТАТЬ:
            - "Сколько заказов у менеджера Иван?"
            - "Какой товар самый популярный?"
            - "Средняя сумма заказа в этом месяце"
            - "Покажи все заказы больше 100000 тенге"
            - "Сравни продажи разных менеджеров"
            - "Какие товары продаются хуже всего?"
            - "Найди заказы клиента XYZ"
            - "Статистика по дням недели"
            - "Прогноз продаж"
            - И любые другие аналитические вопросы!
            
            ФОРМАТ ОТВЕТА:
            - КРАТКО и по делу - максимум 3-5 предложений
            - Сразу давай конкретные цифры
            - Если данных недостаточно - сразу скажи что нужно
            - Никаких длинных объяснений и "воды"
            - Используй эмодзи только для ключевых моментов
            - Если есть несколько вариантов интерпретации - выбери самый логичный
            
            ВАЖНО: Отвечай максимально кратко и конкретно. Пользователю нужен быстрый ответ, а не лекция.
            """
            
            # Получаем ответ от ChatGPT
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Ты эксперт по анализу данных продаж и финансов. Отвечай на русском языке, давай конкретные рекомендации и инсайты."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=2000,
                temperature=0.7
            )
            
            return response.choices[0].message.content
            
        except Exception as e:
            return f"Ошибка при обработке запроса: {str(e)}"
    
    def _is_edit_command(self, user_query: str) -> bool:
        """
        Определяет, является ли запрос командой редактирования
        
        Args:
            user_query (str): Запрос пользователя
            
        Returns:
            bool: True если это команда редактирования
        """
        edit_keywords = [
            'изменить', 'изменение', 'поменять', 'исправить', 'обновить',
            'заказ', 'заказе', 'котел', 'котлы', 'количество', 'цена', 'цену',
            'вместо', 'взял', 'заказал', 'нужно', 'надо', 'менеджер', 'менеджера',
            'поставь', 'назначь', 'установи', 'смени'
        ]
        
        query_lower = user_query.lower()
        return any(keyword in query_lower for keyword in edit_keywords)
    
    def execute_command(self, excel_file_path: str, command: str) -> str:
        """
        Выполняет любую команду пользователя
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            command (str): Команда пользователя
            
        Returns:
            str: Результат выполнения команды
        """
        if not self.client:
            return "❌ ChatGPT AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            # Читаем данные из Excel
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            
            # Создаем универсальный промпт для выполнения команд
            prompt = f"""
            Ты - умный помощник для работы с данными продаж. Пользователь дал команду: "{command}"
            
            ДОСТУПНЫЕ ДАННЫЕ:
            Колонки: {list(sales_df.columns)}
            Количество записей: {len(sales_df)}
            
            ПЕРВЫЕ 5 ЗАПИСЕЙ:
            {sales_df.head(5).to_dict('records')}
            
            ТИПЫ КОМАНД, КОТОРЫЕ ТЫ МОЖЕШЬ ВЫПОЛНИТЬ:
            
            1. РЕДАКТИРОВАНИЕ ДАННЫХ:
            - "изменить менеджера в заказе 1 на Алибек"
            - "поменять количество в заказе 2 на 3"
            - "обновить цену в заказе 3 на 50000"
            - "назначить менеджера Айдана в заказе 4"
            
            2. ПОИСК И ФИЛЬТРАЦИЯ:
            - "найди все заказы менеджера Алибек"
            - "покажи заказы больше 100000 тенге"
            - "найди заказы клиента Иван"
            - "покажи все заказы за сегодня"
            
            3. АНАЛИЗ И СТАТИСТИКА:
            - "сколько заказов у каждого менеджера"
            - "какая средняя сумма заказа"
            - "какой товар самый популярный"
            - "покажи топ-5 менеджеров"
            
            4. ОТЧЕТЫ:
            - "создай отчет за эту неделю"
            - "покажи статистику по месяцам"
            - "сравни продажи разных менеджеров"
            
            ПРАВИЛА ВЫПОЛНЕНИЯ:
            - Если это команда редактирования - дай точные инструкции в формате EDIT_INSTRUCTIONS
            - Если это поиск - найди и покажи конкретные записи
            - Если это анализ - проведи расчеты и дай результат
            - Если это отчет - создай структурированный отчет
            
            ФОРМАТ ОТВЕТА:
            - Для редактирования: EDIT_INSTRUCTIONS: [инструкции]
            - Для поиска: НАЙДЕНО: [результаты]
            - Для анализа: АНАЛИЗ: [результаты]
            - Для отчетов: ОТЧЕТ: [отчет]
            
            ВАЖНО: Отвечай кратко и по делу. Если команда неясна - уточни что именно нужно сделать.
            """
            
            # Получаем ответ от Gemini
            response = self.model.generate_content(prompt)
            result = response.text
            
            # Если это команда редактирования, применяем изменения
            if "EDIT_INSTRUCTIONS:" in result:
                return self._apply_edit_instructions(excel_file_path, sales_df, result)
            
            return result
            
        except Exception as e:
            return f"Ошибка при выполнении команды: {str(e)}"
    
    def edit_excel_data(self, excel_file_path: str, edit_request: str) -> str:
        """
        Редактирует данные в Excel файле на основе запроса пользователя
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            edit_request (str): Текстовый запрос на изменение (например: "заказ номер #2 вместо двух котлов взял один")
            
        Returns:
            str: Результат операции
        """
        print(f"🔍 CHATGPT DEBUG: edit_excel_data вызван")
        print(f"🔍 CHATGPT DEBUG: excel_file_path = {excel_file_path}")
        print(f"🔍 CHATGPT DEBUG: edit_request = '{edit_request}'")
        print(f"🔍 CHATGPT DEBUG: self.client = {self.client}")
        
        if not self.client:
            print(f"🔍 CHATGPT DEBUG: Клиент недоступен")
            return "❌ ChatGPT AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            print(f"🔍 CHATGPT DEBUG: Читаем Excel файл...")
            # Читаем данные из Excel с использованием кэша
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            print(f"🔍 CHATGPT DEBUG: Прочитано {len(sales_df)} строк")
            print(f"🔍 CHATGPT DEBUG: Колонки: {sales_df.columns.tolist()}")
            
            print(f"🔍 CHATGPT DEBUG: Создаем промпт...")
            # Создаем промпт для понимания изменений
            prompt = self._create_edit_prompt(sales_df, edit_request)
            print(f"🔍 CHATGPT DEBUG: Промпт создан, длина: {len(prompt)}")
            
            print(f"🔍 CHATGPT DEBUG: Отправляем запрос к ChatGPT...")
            # Получаем инструкции от ChatGPT
            response = self.client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": "Ты эксперт по анализу и редактированию данных. Отвечай на русском языке, давай точные инструкции для изменения данных."},
                    {"role": "user", "content": prompt}
                ],
                max_tokens=1500,
                temperature=0.3
            )
            edit_instructions = response.choices[0].message.content
            print(f"🔍 CHATGPT DEBUG: Получен ответ от ChatGPT: {edit_instructions[:200]}...")
            
            print(f"🔍 CHATGPT DEBUG: Применяем инструкции...")
            # Парсим инструкции и применяем изменения
            result = self._apply_edit_instructions(excel_file_path, sales_df, edit_instructions)
            print(f"🔍 CHATGPT DEBUG: Результат применения: {result[:100]}...")
            
            return result
            
        except Exception as e:
            print(f"🔍 CHATGPT DEBUG: Ошибка в edit_excel_data: {str(e)}")
            import traceback
            print(f"🔍 CHATGPT DEBUG: Полный traceback:")
            traceback.print_exc()
            return f"Ошибка при редактировании данных: {str(e)}"
    
    def _create_edit_prompt(self, sales_df: pd.DataFrame, edit_request: str) -> str:
        """Создает промпт для понимания изменений"""
        # Показываем первые несколько записей для контекста
        sample_data = sales_df.head(5).to_dict('records')
        
        # Получаем список колонок для лучшего понимания структуры
        columns_info = list(sales_df.columns)
        
        prompt = f"""
        Пользователь просит внести изменения в данные о продажах. Вот текущие данные:

        ЗАПРОС ПОЛЬЗОВАТЕЛЯ: "{edit_request}"

        СТРУКТУРА ДАННЫХ (колонки):
        {columns_info}

        ТЕКУЩИЕ ДАННЫЕ (первые 5 записей):
        {sample_data}

        ВАЖНО! Структура данных на английском языке:
        - Номер заказа: 'order' 
        - Дата: 'date'
        - Клиент: 'client'
        - Название котла/товара: 'name_boiler'
        - Количество: 'quantity'
        - Цена: 'price'
        - Покупка: 'purchase'
        - Оплата: 'payment'
        - Доставка: 'delivery'
        - Аксессуары: 'accessories'
        - Основной менеджер: 'manager'
        - Другие менеджеры могут быть в колонках: manager_1, manager_2, manager_3, manager_4, manager_5, manager_6, manager_7
        - ROP сотрудник: 'employee ROP'
        - Разработчик: 'developer'
        - Помощники: 'assistant', 'assistant 2'
        - Менеджер поставщика: 'supplier manager'

        ПРАВИЛА ПОИСКА И СОЗДАНИЯ:
        - Если упоминается "заказ номер #X" или "заказе X", ищи по колонке 'order'
        - Если упоминается товар "котлы", "котел", ищи по колонке 'name_boiler'
        - Если упоминается имя человека, ищи в колонках менеджеров (manager, manager_1, manager_2, и т.д.)
        - Если упоминается количество, работай с колонкой 'quantity'
        - Если упоминается цена, работай с колонкой 'price'
        
        КОМАНДЫ СОЗДАНИЯ НОВЫХ ЗАКАЗОВ:
        - Если запрос содержит "Создай новый заказ" или "создай новый заказ", создавай новую запись
        - Извлекай данные: номер заказа, дату, товар, менеджера, доставку
        - Для даты используй формат YYYY-MM-DD (например, 2025-09-12)

        Пожалуйста, проанализируй запрос и предоставь точные инструкции в следующем формате:

        ДЛЯ РЕДАКТИРОВАНИЯ СУЩЕСТВУЮЩИХ ЗАПИСЕЙ:
        EDIT_INSTRUCTIONS:
        - Найти запись: [условие поиска, например: order == 2]
        - Изменить поле: [название_поля] = [новое_значение]
        - Пересчитать: [название_поля] = [формула_пересчета]

        ДЛЯ СОЗДАНИЯ НОВЫХ ЗАКАЗОВ:
        EDIT_INSTRUCTIONS:
        - Создать новый заказ: [номер_заказа]
        - Установить поле: [название_поля] = [значение]
        - Установить поле: [название_поля] = [значение]
        - ... (для всех полей)

        ПРИМЕРЫ:
        - Для "заказ номер #2": order == 2
        - Для "количество котлов": quantity
        - Для "менеджер Иван": найти в manager или manager_1, manager_2, и т.д.
        - Для "Создай новый заказ номер 19": создать новую запись с order = 19

        Отвечай только инструкциями в указанном формате, без дополнительных объяснений.
        """
        return prompt
    
    def _apply_edit_instructions(self, excel_file_path: str, sales_df: pd.DataFrame, edit_instructions: str) -> str:
        """Применяет инструкции редактирования к Excel файлу"""
        try:
            # Парсим инструкции
            instructions = self._parse_edit_instructions(edit_instructions)
            
            print(f"🔍 DEBUG: Найдено инструкций: {len(instructions)}")
            
            if not instructions:
                return "❌ Не удалось понять инструкции для изменения"
            
            # Применяем изменения к DataFrame
            modified_df = sales_df.copy()
            results = []
            
            for instruction in instructions:
                if instruction['action'] == 'create_new_order':
                    print("🔍 DEBUG: Входим в блок создания заказа")
                    # Создаем новый заказ
                    new_order_data = instruction['data']
                    order_num = new_order_data.get('order')
                    print(f"🔍 DEBUG: order_num = {order_num}")
                    
                    # Если номер заказа 'auto', находим следующий доступный номер
                    if order_num == 'auto':
                        if 'order' in modified_df.columns:
                            max_order = modified_df['order'].max() if not modified_df.empty else 0
                            order_num = int(max_order) + 1
                        else:
                            order_num = 1
                    
                    # Создаем новую строку
                    new_row = {}
                    for field, value in new_order_data.items():
                        if field != 'order':
                            new_row[field] = value
                    
                    # Устанавливаем номер заказа
                    new_row['order'] = order_num
                    
                    # Добавляем новую строку в DataFrame
                    new_row_df = pd.DataFrame([new_row])
                    modified_df = pd.concat([modified_df, new_row_df], ignore_index=True)
                    
                    print(f"✅ Создан новый заказ #{order_num}")
                    
                elif instruction['action'] == 'find_and_edit':
                    # Находим строку по условию
                    condition = instruction['condition']
                    field = instruction['field']
                    new_value = instruction['new_value']
                    
                    # Парсим условие поиска
                    if isinstance(condition, str):
                        # Если условие в виде строки, парсим его
                        if 'order ==' in condition:
                            # Извлекаем номер заказа
                            order_num = int(condition.split('order ==')[1].strip())
                            mask = modified_df['order'] == order_num
                        else:
                            # Другие условия пока не поддерживаются
                            print(f"⚠️ Неподдерживаемое условие: {condition}")
                            continue
                    else:
                        # Если условие уже в виде словаря
                        mask = self._apply_condition(modified_df, condition)
                    
                    if mask.any():
                        # Изменяем значение
                        modified_df.loc[mask, field] = new_value
                        
                        # Если изменили количество, пересчитываем цену (если нужно)
                        if field in ['quantity', 'количество', 'кол-во']:
                            # Пересчитываем цену пропорционально, если есть базовая цена
                            if 'price' in modified_df.columns:
                                # Если есть базовая цена, пересчитываем общую стоимость
                                original_qty = sales_df.loc[mask, field].iloc[0] if len(sales_df.loc[mask, field]) > 0 else 1
                                if original_qty > 0:
                                    price_per_unit = sales_df.loc[mask, 'price'].iloc[0] / original_qty if 'price' in sales_df.columns else 0
                                    modified_df.loc[mask, 'price'] = new_value * price_per_unit
                        
                        # Сохраняем изменения в Excel
                        self._save_excel_changes(excel_file_path, modified_df)
                        
                        return f"✅ Изменения успешно внесены!\n\nНайдена запись: {condition}\nИзменено поле '{field}' на значение: {new_value}"
                    else:
                        return f"❌ Не найдена запись по условию: {condition}"
                
                elif instruction['action'] == 'create_new_order':
                    # Создаем новый заказ
                    new_order_data = instruction['data']
                    order_num = new_order_data.get('order')
                    
                    # Если номер заказа 'auto', находим следующий доступный номер
                    if order_num == 'auto':
                        if 'order' in modified_df.columns:
                            max_order = modified_df['order'].max() if not modified_df.empty else 0
                            order_num = int(max_order) + 1
                        else:
                            order_num = 1
                    
                    # Создаем новую строку
                    new_row = {}
                    for field, value in new_order_data.items():
                        if field != 'order':
                            new_row[field] = value
                    
                    # Устанавливаем номер заказа
                    new_row['order'] = order_num
                    print(f"🔍 DEBUG: new_row создан: {new_row}")
                    
                    # Добавляем новую строку в DataFrame
                    try:
                        new_row_df = pd.DataFrame([new_row])
                        modified_df = pd.concat([modified_df, new_row_df], ignore_index=True)
                        print(f"🔍 DEBUG: Строка добавлена в DataFrame")
                    except Exception as e:
                        print(f"❌ Ошибка при добавлении строки: {e}")
                        # Не используем continue, продолжаем выполнение
                    
                    # Создаем простое сообщение о заказе
                    product = new_order_data.get('product', 'не указан')
                    manager = new_order_data.get('manager', 'не указан')
                    delivery = new_order_data.get('delivery', 'не указана')
                    quantity = new_order_data.get('quantity', 1)
                    
                    print(f"🔍 DEBUG: Доходим до создания сообщения о заказе")
                    
                    result_msg = f"✅ Новый заказ #{order_num} создан! Товар: {product}, Менеджер: {manager}, Доставка: {delivery}"
                    
                    # Добавляем результат в список
                    results.append(result_msg)
                    print(f"✅ Создан новый заказ #{order_num}")
                    print(f"🔍 DEBUG: Результат добавлен в список, всего результатов: {len(results)}")
                    
                    # Проверяем, что мы дошли до этого места
                    print(f"🔍 DEBUG: Доходим до конца блока создания заказа")
            
            # Сохраняем все изменения в Excel один раз в конце
            if not modified_df.equals(sales_df):
                self._save_excel_changes(excel_file_path, modified_df)
                print("✅ Все изменения сохранены в Excel файл")
            
            # Возвращаем результат
            if results:
                return "\n".join(results)
            else:
                return "❌ Не удалось применить изменения"
            
        except Exception as e:
            return f"Ошибка при применении изменений: {str(e)}"
    
    def _parse_edit_instructions(self, instructions_text: str) -> List[Dict[str, Any]]:
        """Парсит инструкции редактирования"""
        instructions = []
        
        # Ищем блок EDIT_INSTRUCTIONS
        if "EDIT_INSTRUCTIONS:" in instructions_text:
            instructions_section = instructions_text.split("EDIT_INSTRUCTIONS:")[1]
        else:
            instructions_section = instructions_text
        
        lines = instructions_section.strip().split('\n')
        
        current_condition = None
        
        current_order_data = {}
        is_creating_new_order = False
        
        for line in lines:
            line = line.strip()
            if line.startswith('- Создать новый заказ:') or line.startswith('Создать новый заказ:'):
                order_num = line.replace('- Создать новый заказ:', '').replace('Создать новый заказ:', '').strip()
                # Извлекаем только число из строки типа "[номер_заказа] = 6"
                if '=' in order_num:
                    order_num = order_num.split('=')[1].strip()
                # Убираем скобки и извлекаем число
                order_num = order_num.replace('[', '').replace(']', '').strip()
                if order_num.isdigit():
                    current_order_data['order'] = int(order_num)
                else:
                    # Если не можем извлечь число, используем следующий доступный номер
                    current_order_data['order'] = 'auto'
                is_creating_new_order = True
            elif line.startswith('- Установить поле:') or line.startswith('Установить поле:'):
                field_part = line.replace('- Установить поле:', '').replace('Установить поле:', '').strip()
                if '=' in field_part:
                    field, new_value = field_part.split('=', 1)
                    field = field.strip()
                    new_value = new_value.strip()
                    
                    # Преобразуем значение в правильный тип
                    if new_value.isdigit():
                        new_value = int(new_value)
                    elif new_value.replace('.', '').isdigit():
                        new_value = float(new_value)
                    elif new_value.startswith('"') and new_value.endswith('"'):
                        new_value = new_value[1:-1]
                    elif new_value.startswith("'") and new_value.endswith("'"):
                        new_value = new_value[1:-1]
                    
                    if is_creating_new_order:
                        current_order_data[field] = new_value
                    else:
                        instructions.append({
                            'action': 'find_and_edit',
                            'condition': current_condition or '',
                            'field': field,
                            'new_value': new_value
                        })
            elif line.startswith('- Найти запись:') or line.startswith('Найти запись:'):
                condition = line.replace('- Найти запись:', '').replace('Найти запись:', '').strip()
                current_condition = condition
                is_creating_new_order = False
            elif line.startswith('- Изменить поле:') or line.startswith('Изменить поле:'):
                field_part = line.replace('- Изменить поле:', '').replace('Изменить поле:', '').strip()
                if '=' in field_part:
                    field, new_value = field_part.split('=', 1)
                    field = field.strip()
                    new_value = new_value.strip()
                    
                    # Преобразуем значение в правильный тип
                    if new_value.isdigit():
                        new_value = int(new_value)
                    elif new_value.replace('.', '').isdigit():
                        new_value = float(new_value)
                    elif new_value.startswith('"') and new_value.endswith('"'):
                        new_value = new_value[1:-1]
                    elif new_value.startswith("'") and new_value.endswith("'"):
                        new_value = new_value[1:-1]
                    
                    instructions.append({
                        'action': 'find_and_edit',
                        'condition': current_condition or '',
                        'field': field,
                        'new_value': new_value
                    })
        
        # Если создаем новый заказ, добавляем инструкцию
        if is_creating_new_order and current_order_data:
            instructions.append({
                'action': 'create_new_order',
                'data': current_order_data
            })
        
        # Если инструкции не найдены, пытаемся извлечь из текста напрямую
        if not instructions or not any(inst.get('action') == 'create_new_order' for inst in instructions):
            # Проверяем, является ли это командой создания нового заказа
            if ('Создай новый заказ' in instructions_text or 'создай новый заказ' in instructions_text or 
                'Добавь новый заказ' in instructions_text or 'добавь новый заказ' in instructions_text):
                # Парсим данные нового заказа
                import re
                from datetime import datetime
                
                order_data = {}
                
                # Извлекаем номер заказа
                order_match = re.search(r'номер\s+(\d+)', instructions_text)
                if order_match:
                    order_data['order'] = int(order_match.group(1))
                
                # Извлекаем дату
                date_match = re.search(r'(\d{1,2})\s+сентября', instructions_text)
                if date_match:
                    day = int(date_match.group(1))
                    order_data['date'] = f"2025-09-{day:02d}"
                
                    # Извлекаем товар (поддерживаем разные варианты)
                    product_match = re.search(r'(alseit_\d+|баусит\s+\d+|balseit\s*_\d+)', instructions_text.lower())
                if product_match:
                        product = product_match.group(1)
                        # Нормализуем название товара
                        if 'баусит' in product:
                            product = product.replace('баусит', 'balseit')
                        # Убираем пробелы вокруг подчеркивания
                        product = product.replace(' ', '')
                        order_data['product'] = product
                
                # Извлекаем менеджера
                manager_match = re.search(r'менеджер\s+(\w+)', instructions_text.lower())
                if manager_match:
                    order_data['manager'] = manager_match.group(1)
                
                # Извлекаем количество
                quantity_match = re.search(r'(\d+)\s+штук', instructions_text.lower())
                if quantity_match:
                    order_data['quantity'] = int(quantity_match.group(1))
                else:
                    order_data['quantity'] = 1  # По умолчанию 1 штука
                    
                # Извлекаем доставку
                delivery_match = re.search(r'доставка\s+в\s+(\w+)', instructions_text.lower())
                if delivery_match:
                    order_data['delivery'] = delivery_match.group(1)
                else:
                    # Пробуем без "в"
                    delivery_match = re.search(r'доставка\s+(\w+)', instructions_text.lower())
                    if delivery_match:
                        order_data['delivery'] = delivery_match.group(1)
                
                if order_data:
                    instructions.append({
                        'action': 'create_new_order',
                        'data': order_data
                    })
            else:
                # Ищем паттерны типа "order == 1" и "manager = Айдана"
                import re
                order_match = re.search(r'order\s*==\s*(\d+)', instructions_text)
                field_match = re.search(r'(\w+)\s*=\s*([^,\n]+)', instructions_text)
            
            if order_match and field_match:
                order_num = int(order_match.group(1))
                field = field_match.group(1).strip()
                new_value = field_match.group(2).strip()
                
                # Преобразуем значение
                if new_value.isdigit():
                    new_value = int(new_value)
                elif new_value.replace('.', '').isdigit():
                    new_value = float(new_value)
                elif new_value.startswith('"') and new_value.endswith('"'):
                    new_value = new_value[1:-1]
                elif new_value.startswith("'") and new_value.endswith("'"):
                    new_value = new_value[1:-1]
                
                instructions.append({
                    'action': 'find_and_edit',
                    'condition': f'order == {order_num}',
                    'field': field,
                    'new_value': new_value
                })
        
        return instructions
    
    def _apply_condition(self, df: pd.DataFrame, condition: Dict[str, Any]) -> pd.Series:
        """
        Применяет условие поиска к DataFrame
        
        Args:
            df: DataFrame для поиска
            condition: Словарь с условием поиска
            
        Returns:
            pandas.Series: Маска для фильтрации строк
        """
        try:
            # Начинаем с True для всех строк
            mask = pd.Series([True] * len(df), index=df.index)
            
            # Применяем каждое условие
            for field, value in condition.items():
                if field in df.columns:
                    if isinstance(value, str):
                        # Для строковых значений ищем точное совпадение или частичное
                        if value.startswith('*') and value.endswith('*'):
                            # Частичное совпадение
                            search_value = value[1:-1]
                            mask &= df[field].astype(str).str.contains(search_value, case=False, na=False)
                        else:
                            # Точное совпадение
                            mask &= df[field].astype(str).str.lower() == value.lower()
                    else:
                        # Для числовых значений
                        mask &= df[field] == value
                else:
                    print(f"⚠️ Поле '{field}' не найдено в DataFrame")
            
            return mask
            
        except Exception as e:
            print(f"❌ Ошибка применения условия: {e}")
            return pd.Series([False] * len(df), index=df.index)
    
    def _save_excel_changes(self, excel_file_path: str, modified_df: pd.DataFrame) -> None:
        """Сохраняет изменения в Excel файл"""
        try:
            # Загружаем существующий файл
            wb = load_workbook(excel_file_path)
            
            # Обновляем лист 'продажи'
            if 'продажи' in wb.sheetnames:
                ws = wb['продажи']
            else:
                ws = wb.active
            
            # Очищаем лист и записываем новые данные
            ws.delete_rows(1, ws.max_row)
            
            # Записываем заголовки
            for col, header in enumerate(modified_df.columns, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Записываем данные
            for row_idx, row_data in enumerate(modified_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Сохраняем файл
            wb.save(excel_file_path)
            
        except Exception as e:
            print(f"Ошибка при сохранении Excel файла: {e}")
            raise
