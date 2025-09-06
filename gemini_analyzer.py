import os
import pandas as pd
import google.generativeai as genai
from dotenv import load_dotenv
from openpyxl import load_workbook
from utils import load_excel_with_cache, find_column, clean_numeric_data

# Загружаем переменные окружения
load_dotenv()

class GeminiAnalyzer:
    def __init__(self):
        """Инициализация Gemini AI анализатора"""
        self.api_key = os.getenv("GEMINI_API_KEY")
        self.model = None
        
        if self.api_key:
            # Настройка Gemini только если API ключ доступен
            genai.configure(api_key=self.api_key)
            self.model = genai.GenerativeModel('gemini-1.5-flash')
    
    def analyze_sales_data(self, excel_file_path):
        """
        Анализирует данные продаж из Excel файла с помощью Gemini AI
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            
        Returns:
            str: Анализ данных в текстовом формате
        """
        if not self.model:
            return "❌ Gemini AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            # Читаем данные из Excel с использованием кэша
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            salary_df = load_excel_with_cache(excel_file_path, 'зарплата')
            
            # Подготавливаем данные для анализа
            analysis_data = self._prepare_data_for_analysis(sales_df, salary_df)
            
            # Создаем промпт для Gemini
            prompt = self._create_analysis_prompt(analysis_data)
            
            # Получаем анализ от Gemini
            response = self.model.generate_content(prompt)
            
            return response.text
            
        except Exception as e:
            return f"Ошибка при анализе данных: {str(e)}"
    
    def _prepare_data_for_analysis(self, sales_df, salary_df):
        """Подготавливает данные для анализа"""
        
        # Определяем колонку с ценами/суммами
        price_column = None
        possible_price_columns = ['price', 'сумма', 'стоимость', 'total', 'amount']
        
        for col in possible_price_columns:
            if col in sales_df.columns:
                price_column = col
                break
        
        # Определяем колонку с основным менеджером
        manager_column = None
        possible_manager_columns = ['manager', 'менеджер']
        
        for col in possible_manager_columns:
            if col in sales_df.columns:
                manager_column = col
                break
        
        # Определяем колонку с ROP
        rop_column = None
        possible_rop_columns = ['employee ROP', 'rop', 'РОП']
        
        for col in possible_rop_columns:
            if col in sales_df.columns:
                rop_column = col
                break
        
        # Основная статистика по продажам
        sales_stats = {
            'total_sales': pd.to_numeric(sales_df[price_column], errors='coerce').sum() if price_column else 0,
            'avg_sale': pd.to_numeric(sales_df[price_column], errors='coerce').mean() if price_column else 0,
            'total_transactions': len(sales_df),
            'unique_managers': sales_df[manager_column].nunique() if manager_column else 0,
            'unique_rop': sales_df[rop_column].nunique() if rop_column else 0,
            'price_column_used': price_column,
            'manager_column_used': manager_column,
            'rop_column_used': rop_column
        }
        
        # Статистика по менеджерам
        if manager_column and price_column:
            try:
                # Убираем NaN значения для корректного группирования
                valid_data = sales_df.dropna(subset=[manager_column])
                valid_data[price_column] = pd.to_numeric(valid_data[price_column], errors='coerce')
                manager_stats = valid_data.groupby(manager_column)[price_column].agg(['sum', 'count', 'mean']).round(2)
                sales_stats['manager_performance'] = manager_stats.to_dict('index')
            except Exception as e:
                sales_stats['manager_performance'] = {}
        
        # Статистика по ROP сотрудникам
        if rop_column and price_column:
            try:
                valid_data = sales_df.dropna(subset=[rop_column])
                valid_data[price_column] = pd.to_numeric(valid_data[price_column], errors='coerce')
                rop_stats = valid_data.groupby(rop_column)[price_column].agg(['sum', 'count', 'mean']).round(2)
                sales_stats['rop_performance'] = rop_stats.to_dict('index')
            except Exception as e:
                sales_stats['rop_performance'] = {}
        
        # Статистика по зарплатам (если есть лист зарплата)
        salary_stats = {'note': 'Данные по зарплатам не анализируются для листа продаж'}
        if salary_df is not None and not salary_df.empty:
            # Определяем колонки зарплат
            salary_column = None
            bonus_columns = []
            
            for col in salary_df.columns:
                if 'оклад' in col.lower() or 'salary' in col.lower():
                    salary_column = col
                elif 'бонус' in col.lower() or 'bonus' in col.lower():
                    bonus_columns.append(col)
            
            if salary_column:
                salary_stats = {
                    'total_salary': pd.to_numeric(salary_df[salary_column], errors='coerce').sum(),
                    'avg_salary': pd.to_numeric(salary_df[salary_column], errors='coerce').mean(),
                    'total_bonuses': sum(pd.to_numeric(salary_df[col], errors='coerce').sum() for col in bonus_columns),
                    'salary_column_used': salary_column,
                    'bonus_columns_used': bonus_columns
                }
        
        return {
            'sales': sales_stats,
            'salary': salary_stats,
            'raw_sales_data': sales_df.head(10).to_dict('records'),  # Первые 10 записей
            'raw_salary_data': salary_df.head(10).to_dict('records') if salary_df is not None else []
        }
    
    def _create_analysis_prompt(self, data):
        """Создает промпт для анализа данных"""
        prompt = f"""
        Проанализируй следующие данные о продажах и зарплатах компании:

        ОСНОВНАЯ СТАТИСТИКА ПРОДАЖ:
        - Общая сумма продаж: {data['sales']['total_sales']:,.0f} тенге
        - Средняя сумма продажи: {data['sales']['avg_sale']:,.0f} тенге
        - Общее количество транзакций: {data['sales']['total_transactions']}
        - Количество менеджеров: {data['sales']['unique_managers']}
        - Количество ROP сотрудников: {data['sales']['unique_rop']}

        СТАТИСТИКА ПО ЗАРПЛАТАМ:
        {self._format_salary_stats(data['salary'])}

        ПЕРВЫЕ 10 ЗАПИСЕЙ ПРОДАЖ:
        {data['raw_sales_data']}

        ПЕРВЫЕ 10 ЗАПИСЕЙ ЗАРПЛАТ:
        {data['raw_salary_data']}

        Проведи КРАТКИЙ анализ и предоставь:
        1. Ключевые показатели (2-3 главные цифры)
        2. Лучшие и худшие результаты
        3. Главный вывод и 1-2 рекомендации

        ВАЖНО: Отвечай кратко, без лишних объяснений. Максимум 5-7 предложений.
        """
        return prompt
    
    def _format_salary_stats(self, salary_data):
        """Форматирует статистику по зарплатам"""
        if isinstance(salary_data, dict) and 'note' in salary_data:
            return salary_data['note']
        
        if isinstance(salary_data, dict):
            result = ""
            if 'total_salary' in salary_data:
                result += f"- Общая сумма окладов: {salary_data['total_salary']:,.0f} тенге\n"
            if 'avg_salary' in salary_data:
                result += f"- Средний оклад: {salary_data['avg_salary']:,.0f} тенге\n"
            if 'total_bonuses' in salary_data:
                result += f"- Общая сумма бонусов: {salary_data['total_bonuses']:,.0f} тенге\n"
            
            if not result:
                return "- Данные по зарплатам недоступны для анализа"
            
            return result.strip()
        
        return "- Данные по зарплатам недоступны"
    
    def get_quick_insights(self, excel_file_path):
        """
        Получает быстрые инсайты по данным
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            
        Returns:
            str: Краткие инсайты
        """
        try:
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            
            # Определяем колонку с ценами/суммами
            price_column = None
            possible_price_columns = ['price', 'сумма', 'стоимость', 'total', 'amount']
            
            for col in possible_price_columns:
                if col in sales_df.columns:
                    price_column = col
                    break
            
            if price_column is None:
                return f"❌ Не найдена колонка с ценами. Доступные колонки: {list(sales_df.columns)}"
            
            # Быстрые расчеты
            # Убираем NaN значения и преобразуем в числовой формат
            price_data = pd.to_numeric(sales_df[price_column], errors='coerce').dropna()
            
            total_sales = price_data.sum()
            avg_sale = price_data.mean()
            max_sale = price_data.max()
            min_sale = price_data.min()
            
            # Дополнительная информация по товарам
            product_info = ""
            if 'name_boiler' in sales_df.columns:
                product_counts = sales_df['name_boiler'].value_counts().head(3)
                product_info = f"\n\n🏆 ТОП-3 ТОВАРА:\n"
                for i, (product, count) in enumerate(product_counts.items(), 1):
                    product_info += f"• {i}. {product}: {count} шт.\n"
            
            # Информация по менеджерам
            manager_info = ""
            manager_columns = [col for col in sales_df.columns if 'manager' in col.lower()]
            if manager_columns:
                # Берем основной столбец менеджера
                main_manager_col = 'manager' if 'manager' in sales_df.columns else manager_columns[0]
                manager_counts = sales_df[main_manager_col].value_counts().head(3)
                manager_info = f"\n\n👥 ТОП-3 МЕНЕДЖЕРА:\n"
                for i, (manager, count) in enumerate(manager_counts.items(), 1):
                    if pd.notna(manager) and str(manager) != 'nan':
                        manager_info += f"• {i}. {manager}: {count} заказов\n"
            
            insights = f"""
📊 БЫСТРЫЕ ИНСАЙТЫ:
• Общая сумма продаж: {total_sales:,.0f} тенге
• Средняя сумма продажи: {avg_sale:,.0f} тенге
• Максимальная продажа: {max_sale:,.0f} тенге
• Минимальная продажа: {min_sale:,.0f} тенге
• Количество транзакций: {len(sales_df)}
• Колонка с ценами: {price_column}{product_info}{manager_info}
            """
            
            return insights.strip()
            
        except Exception as e:
            return f"Ошибка при получении инсайтов: {str(e)}"
    
    def generate_period_report(self, excel_file_path, period_request):
        """
        Генерирует отчет за определенный период
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            period_request (str): Запрос на отчет (например: "отчет за эту неделю")
            
        Returns:
            str: Отчет за период
        """
        if not self.model:
            return "❌ Gemini AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            # Читаем данные из Excel
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            
            # Создаем специальный промпт для отчетов
            prompt = f"""
            Пользователь запросил: "{period_request}"
            
            Проанализируй данные о продажах и создай подробный отчет.
            
            ДАННЫЕ О ПРОДАЖАХ:
            Колонки в данных: {list(sales_df.columns)}
            Количество записей: {len(sales_df)}
            
            ПЕРВЫЕ 10 ЗАПИСЕЙ:
            {sales_df.head(10).to_dict('records')}
            
            ПОСЛЕДНИЕ 5 ЗАПИСЕЙ:
            {sales_df.tail(5).to_dict('records')}
            
            СТАТИСТИКА:
            - Общее количество заказов: {len(sales_df)}
            """
            
            # Добавляем статистику по ценам, если есть
            price_column = find_column(sales_df, 'price')
            if price_column:
                price_data = clean_numeric_data(sales_df, price_column)
                prompt += f"""
            - Общая сумма продаж: {price_data.sum():,.0f} тенге
            - Средняя сумма заказа: {price_data.mean():,.0f} тенге
            - Максимальная продажа: {price_data.max():,.0f} тенге
            - Минимальная продажа: {price_data.min():,.0f} тенге
                """
            
            # Добавляем информацию по менеджерам
            manager_column = find_column(sales_df, 'manager')
            if manager_column:
                manager_stats = sales_df[manager_column].value_counts().head(5)
                prompt += f"""
            
            ТОП-5 МЕНЕДЖЕРОВ ПО КОЛИЧЕСТВУ ЗАКАЗОВ:
            {manager_stats.to_dict()}
                """
            
            # Добавляем информацию по товарам
            product_column = find_column(sales_df, 'product')
            if product_column:
                product_stats = sales_df[product_column].value_counts().head(5)
                prompt += f"""
            
            ТОП-5 ТОВАРОВ ПО КОЛИЧЕСТВУ ЗАКАЗОВ:
            {product_stats.to_dict()}
                """
            
            prompt += f"""
            
            ЗАДАЧА:
            Создай КРАТКИЙ отчет:
            1. Основные цифры (продажи, заказы, средний чек)
            2. ТОП-3 менеджера и товара
            3. Главный вывод
            
            ВАЖНО: 
            - Максимум 6-8 предложений
            - Конкретные цифры, минимум текста
            - Если нет данных за период - скажи сразу
            """
            
            # Получаем анализ от Gemini
            response = self.model.generate_content(prompt)
            
            return response.text
            
        except Exception as e:
            return f"Ошибка при генерации отчета: {str(e)}"
    
    def universal_query(self, excel_file_path, user_query):
        """
        Универсальный обработчик запросов - может ответить на любой вопрос по данным
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            user_query (str): Любой запрос пользователя
            
        Returns:
            str: Ответ на запрос
        """
        if not self.model:
            return "❌ Gemini AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            # Читаем данные из Excel
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            
            # Создаем универсальный промпт
            prompt = f"""
            Ты - умный аналитик данных. Пользователь задал вопрос о данных продаж.
            
            ЗАПРОС ПОЛЬЗОВАТЕЛЯ: "{user_query}"
            
            СПЕЦИАЛЬНЫЕ ПРАВИЛА:
            - Если спрашивают про "средний чек" - используй колонку 'price' как основную сумму заказа
            - Если спрашивают про "прибыль" - используй колонку 'purchase' 
            - Если нужны даты, а колонки 'date' нет - скажи сразу "Нет данных по датам"
            
            ДАННЫЕ О ПРОДАЖАХ:
            Структура данных (колонки): {list(sales_df.columns)}
            Общее количество записей: {len(sales_df)}
            
            ОБРАЗЕЦ ДАННЫХ (первые 5 записей):
            {sales_df.head(5).to_dict('records')}
            
            ПОСЛЕДНИЕ 3 ЗАПИСИ:
            {sales_df.tail(3).to_dict('records')}
            
            СТАТИСТИЧЕСКАЯ ИНФОРМАЦИЯ:
            """
            
            # Добавляем базовую статистику для всех числовых колонок
            numeric_columns = sales_df.select_dtypes(include=['number']).columns
            for col in numeric_columns:
                if not sales_df[col].isna().all():
                    prompt += f"""
            {col}:
              - Минимум: {sales_df[col].min()}
              - Максимум: {sales_df[col].max()}
              - Среднее: {sales_df[col].mean():.2f}
              - Сумма: {sales_df[col].sum()}
                    """
            
            # Добавляем информацию по категориальным колонкам
            categorical_columns = sales_df.select_dtypes(include=['object', 'string']).columns
            for col in categorical_columns[:5]:  # Ограничиваем до 5 колонок
                if not sales_df[col].isna().all():
                    unique_count = sales_df[col].nunique()
                    if unique_count <= 20:  # Показываем только если уникальных значений не много
                        value_counts = sales_df[col].value_counts().head(5)
                        prompt += f"""
            {col} (уникальных значений: {unique_count}):
              Топ-5 значений: {value_counts.to_dict()}
                        """
                    else:
                        prompt += f"""
            {col}: {unique_count} уникальных значений
                        """
            
            prompt += f"""
            
            ТВОЯ ЗАДАЧА:
            1. Проанализируй запрос пользователя и пойми, что именно он хочет узнать
            2. Используй предоставленные данные для ответа
            3. Если нужно сделать вычисления - сделай их
            4. Если нужно найти конкретные записи - найди их
            5. Если нужно сравнить или проанализировать тренды - проанализируй
            6. Если данных недостаточно для полного ответа - скажи об этом честно
            
            ТИПЫ ЗАПРОСОВ, КОТОРЫЕ ТЫ МОЖЕШЬ ОБРАБОТАТЬ:
            - "Сколько заказов у менеджера Иван?"
            - "Какой товар самый популярный?"
            - "Средняя сумма заказа в этом месяце"
            - "Покажи все заказы больше 100000 тенге"
            - "Сравни продажи разных менеджеров"
            - "Какие товары продаются хуже всего?"
            - "Найди заказы клиента XYZ"
            - "Статистика по дням недели"
            - "Прогноз продаж"
            - И любые другие аналитические вопросы!
            
            ФОРМАТ ОТВЕТА:
            - КРАТКО и по делу - максимум 3-5 предложений
            - Сразу давай конкретные цифры
            - Если данных недостаточно - сразу скажи что нужно
            - Никаких длинных объяснений и "воды"
            - Используй эмодзи только для ключевых моментов
            - Если есть несколько вариантов интерпретации - выбери самый логичный
            
            ВАЖНО: Отвечай максимально кратко и конкретно. Пользователю нужен быстрый ответ, а не лекция.
            """
            
            # Получаем ответ от Gemini
            response = self.model.generate_content(prompt)
            
            return response.text
            
        except Exception as e:
            return f"Ошибка при обработке запроса: {str(e)}"
    
    def edit_excel_data(self, excel_file_path, edit_request):
        """
        Редактирует данные в Excel файле на основе запроса пользователя
        
        Args:
            excel_file_path (str): Путь к Excel файлу
            edit_request (str): Текстовый запрос на изменение (например: "заказ номер #2 вместо двух котлов взял один")
            
        Returns:
            str: Результат операции
        """
        print(f"🔍 GEMINI DEBUG: edit_excel_data вызван")
        print(f"🔍 GEMINI DEBUG: excel_file_path = {excel_file_path}")
        print(f"🔍 GEMINI DEBUG: edit_request = '{edit_request}'")
        print(f"🔍 GEMINI DEBUG: self.model = {self.model}")
        
        if not self.model:
            print(f"🔍 GEMINI DEBUG: Модель недоступна")
            return "❌ Gemini AI недоступен. Проверьте настройки API ключа в .env файле."
        
        try:
            print(f"🔍 GEMINI DEBUG: Читаем Excel файл...")
            # Читаем данные из Excel с использованием кэша
            sales_df = load_excel_with_cache(excel_file_path, 'продажи')
            print(f"🔍 GEMINI DEBUG: Прочитано {len(sales_df)} строк")
            print(f"🔍 GEMINI DEBUG: Колонки: {sales_df.columns.tolist()}")
            
            print(f"🔍 GEMINI DEBUG: Создаем промпт...")
            # Создаем промпт для понимания изменений
            prompt = self._create_edit_prompt(sales_df, edit_request)
            print(f"🔍 GEMINI DEBUG: Промпт создан, длина: {len(prompt)}")
            
            print(f"🔍 GEMINI DEBUG: Отправляем запрос к Gemini...")
            # Получаем инструкции от Gemini
            response = self.model.generate_content(prompt)
            edit_instructions = response.text
            print(f"🔍 GEMINI DEBUG: Получен ответ от Gemini: {edit_instructions[:200]}...")
            
            print(f"🔍 GEMINI DEBUG: Применяем инструкции...")
            # Парсим инструкции и применяем изменения
            result = self._apply_edit_instructions(excel_file_path, sales_df, edit_instructions)
            print(f"🔍 GEMINI DEBUG: Результат применения: {result[:100]}...")
            
            return result
            
        except Exception as e:
            print(f"🔍 GEMINI DEBUG: Ошибка в edit_excel_data: {str(e)}")
            import traceback
            print(f"🔍 GEMINI DEBUG: Полный traceback:")
            traceback.print_exc()
            return f"Ошибка при редактировании данных: {str(e)}"
    
    def _create_edit_prompt(self, sales_df, edit_request):
        """Создает промпт для понимания изменений"""
        # Показываем первые несколько записей для контекста
        sample_data = sales_df.head(5).to_dict('records')
        
        # Получаем список колонок для лучшего понимания структуры
        columns_info = list(sales_df.columns)
        
        prompt = f"""
        Пользователь просит внести изменения в данные о продажах. Вот текущие данные:

        ЗАПРОС ПОЛЬЗОВАТЕЛЯ: "{edit_request}"

        СТРУКТУРА ДАННЫХ (колонки):
        {columns_info}

        ТЕКУЩИЕ ДАННЫЕ (первые 5 записей):
        {sample_data}

        ВАЖНО! Структура данных на английском языке:
        - Номер заказа: 'order' 
        - Дата: 'date'
        - Клиент: 'client'
        - Название котла/товара: 'name_boiler'
        - Количество: 'quantity'
        - Цена: 'price'
        - Покупка: 'purchase'
        - Оплата: 'payment'
        - Доставка: 'delivery'
        - Аксессуары: 'accessories'
        - Основной менеджер: 'manager'
        - Другие менеджеры могут быть в колонках: manager_1, manager_2, manager_3, manager_4, manager_5, manager_6, manager_7
        - ROP сотрудник: 'employee ROP'
        - Разработчик: 'developer'
        - Помощники: 'assistant', 'assistant 2'
        - Менеджер поставщика: 'supplier manager'

        ПРАВИЛА ПОИСКА:
        - Если упоминается "заказ номер #X" или "заказе X", ищи по колонке 'order'
        - Если упоминается товар "котлы", "котел", ищи по колонке 'name_boiler'
        - Если упоминается имя человека, ищи в колонках менеджеров (manager, manager_1, manager_2, и т.д.)
        - Если упоминается количество, работай с колонкой 'quantity'
        - Если упоминается цена, работай с колонкой 'price'

        Пожалуйста, проанализируй запрос и предоставь точные инструкции для изменения в следующем формате:

        EDIT_INSTRUCTIONS:
        - Найти запись: [условие поиска, например: order == 2]
        - Изменить поле: [название_поля] = [новое_значение]
        - Пересчитать: [название_поля] = [формула_пересчета]

        ПРИМЕРЫ:
        - Для "заказ номер #2": order == 2
        - Для "количество котлов": quantity
        - Для "менеджер Иван": найти в manager или manager_1, manager_2, и т.д.

        Отвечай только инструкциями в указанном формате, без дополнительных объяснений.
        """
        return prompt
    
    def _apply_edit_instructions(self, excel_file_path, sales_df, edit_instructions):
        """Применяет инструкции редактирования к Excel файлу"""
        try:
            # Парсим инструкции
            instructions = self._parse_edit_instructions(edit_instructions)
            
            if not instructions:
                return "❌ Не удалось понять инструкции для изменения"
            
            # Применяем изменения к DataFrame
            modified_df = sales_df.copy()
            
            for instruction in instructions:
                if instruction['action'] == 'find_and_edit':
                    # Находим строку по условию
                    condition = instruction['condition']
                    field = instruction['field']
                    new_value = instruction['new_value']
                    
                    # Применяем условие поиска
                    mask = self._apply_condition(modified_df, condition)
                    
                    if mask.any():
                        # Изменяем значение
                        modified_df.loc[mask, field] = new_value
                        
                        # Если изменили количество, пересчитываем цену (если нужно)
                        if field in ['quantity', 'количество', 'кол-во']:
                            # Пересчитываем цену пропорционально, если есть базовая цена
                            if 'price' in modified_df.columns:
                                # Если есть базовая цена, пересчитываем общую стоимость
                                original_qty = sales_df.loc[mask, field].iloc[0] if len(sales_df.loc[mask, field]) > 0 else 1
                                if original_qty > 0:
                                    price_per_unit = sales_df.loc[mask, 'price'].iloc[0] / original_qty if 'price' in sales_df.columns else 0
                                    modified_df.loc[mask, 'price'] = new_value * price_per_unit
                        
                        # Сохраняем изменения в Excel
                        self._save_excel_changes(excel_file_path, modified_df)
                        
                        return f"✅ Изменения успешно внесены!\n\nНайдена запись: {condition}\nИзменено поле '{field}' на значение: {new_value}"
                    else:
                        return f"❌ Не найдена запись по условию: {condition}"
            
            return "❌ Не удалось применить изменения"
            
        except Exception as e:
            return f"Ошибка при применении изменений: {str(e)}"
    
    def _parse_edit_instructions(self, instructions_text):
        """Парсит инструкции редактирования"""
        instructions = []
        
        # Ищем блок EDIT_INSTRUCTIONS
        if "EDIT_INSTRUCTIONS:" in instructions_text:
            instructions_section = instructions_text.split("EDIT_INSTRUCTIONS:")[1]
        else:
            instructions_section = instructions_text
        
        lines = instructions_section.strip().split('\n')
        
        for line in lines:
            line = line.strip()
            if line.startswith('- Найти запись:'):
                condition = line.replace('- Найти запись:', '').strip()
                instructions.append({
                    'action': 'find',
                    'condition': condition
                })
            elif line.startswith('- Изменить поле:'):
                field_part = line.replace('- Изменить поле:', '').strip()
                if '=' in field_part:
                    field, new_value = field_part.split('=', 1)
                    field = field.strip()
                    new_value = new_value.strip()
                    
                    # Преобразуем значение в правильный тип
                    if new_value.isdigit():
                        new_value = int(new_value)
                    elif new_value.replace('.', '').isdigit():
                        new_value = float(new_value)
                    elif new_value.startswith('"') and new_value.endswith('"'):
                        new_value = new_value[1:-1]
                    
                    instructions.append({
                        'action': 'find_and_edit',
                        'condition': instructions[-1]['condition'] if instructions and instructions[-1]['action'] == 'find' else '',
                        'field': field,
                        'new_value': new_value
                    })
        
        return instructions
    
    def _apply_condition(self, df, condition):
        """Применяет условие поиска к DataFrame"""
        try:
            # Простая обработка условий типа "order == 2"
            if '==' in condition:
                field, value = condition.split('==', 1)
                field = field.strip()
                value = value.strip()
                
                # Преобразуем значение
                if value.isdigit():
                    value = int(value)
                elif value.replace('.', '').replace(',', '').isdigit():
                    value = float(value.replace(',', ''))
                elif value.startswith('"') and value.endswith('"'):
                    value = value[1:-1]
                elif value.startswith("'") and value.endswith("'"):
                    value = value[1:-1]
                
                # Прямой поиск по точному названию колонки
                if field in df.columns:
                    return df[field] == value
                
                # Поиск по альтернативным названиям для основных полей
                field_mappings = {
                    'номер_заказа': 'order',
                    'заказ': 'order', 
                    'количество': 'quantity',
                    'кол-во': 'quantity',
                    'товар': 'name_boiler',
                    'наименование': 'name_boiler',
                    'котел': 'name_boiler',
                    'котлы': 'name_boiler',
                    'цена': 'price',
                    'сумма': 'price',  # если нет отдельной колонки суммы
                    'менеджер': 'manager',
                    'клиент': 'client'
                }
                
                if field.lower() in field_mappings:
                    mapped_field = field_mappings[field.lower()]
                    if mapped_field in df.columns:
                        return df[mapped_field] == value
                
                # Если это может быть имя менеджера, ищем во всех колонках менеджеров
                manager_columns = [col for col in df.columns if 'manager' in col.lower()]
                if isinstance(value, str) and len(value) > 2:  # Если значение похоже на имя
                    for col in manager_columns:
                        mask = df[col].astype(str).str.contains(str(value), case=False, na=False)
                        if mask.any():
                            return mask
                
                # Поиск по частичному совпадению названий колонок
                for col in df.columns:
                    if field.lower() in col.lower() or col.lower() in field.lower():
                        return df[col] == value
            
            # Обработка условий поиска по содержимому (например, "manager содержит Иван")
            elif 'содержит' in condition.lower() or 'contains' in condition.lower():
                parts = condition.lower().split('содержит') if 'содержит' in condition.lower() else condition.lower().split('contains')
                if len(parts) == 2:
                    field = parts[0].strip()
                    value = parts[1].strip().strip('"').strip("'")
                    
                    # Ищем в колонках менеджеров
                    manager_columns = [col for col in df.columns if 'manager' in col.lower()]
                    for col in manager_columns:
                        mask = df[col].astype(str).str.contains(str(value), case=False, na=False)
                        if mask.any():
                            return mask
            
            return pd.Series([False] * len(df))
            
        except Exception as e:
            print(f"Ошибка при применении условия: {e}")
            return pd.Series([False] * len(df))
    
    def _save_excel_changes(self, excel_file_path, modified_df):
        """Сохраняет изменения в Excel файл"""
        try:
            # Загружаем существующий файл
            wb = load_workbook(excel_file_path)
            
            # Обновляем лист 'продажи'
            if 'продажи' in wb.sheetnames:
                ws = wb['продажи']
            else:
                ws = wb.active
            
            # Очищаем лист и записываем новые данные
            ws.delete_rows(1, ws.max_row)
            
            # Записываем заголовки
            for col, header in enumerate(modified_df.columns, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Записываем данные
            for row_idx, row_data in enumerate(modified_df.itertuples(index=False), 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Сохраняем файл
            wb.save(excel_file_path)
            
        except Exception as e:
            print(f"Ошибка при сохранении Excel файла: {e}")
            raise
